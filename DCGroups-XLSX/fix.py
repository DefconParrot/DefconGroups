import pandas as pd
import requests
from urllib.parse import urlparse
import matplotlib.pyplot as plt
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import re
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
import os
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import sys
from collections import defaultdict
import time

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def is_valid_url(url):
    try:
        result = urlparse(url)
        return all([result.scheme, result.netloc])
    except:
        return False

def check_link(url, max_retries=3, delay=5):
    logger.info(f"Checking link: {url}")
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept-Language': 'en-US,en;q=0.5',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0'
    }
    for attempt in range(max_retries):
        try:
            response = requests.get(url, timeout=30, allow_redirects=True, headers=headers)
            final_url = response.url
            if response.status_code < 400:
                if final_url != url:
                    logger.info(f"Redirect: {url} -> {final_url}")
                    return ('redirect', final_url)
                else:
                    logger.info(f"Valid link: {url}")
                    return ('valid', url)
            if 'instagram.com' in url or 'reddit.com' in url:
                if 'Page Not Found' not in response.text and 'page not found' not in response.text.lower():
                    logger.info(f"Valid link (special case): {url}")
                    return ('valid', url)
            logger.warning(f"Broken link: {url} (Status code: {response.status_code})")
            return ('broken', url)
        except requests.exceptions.RequestException as e:
            logger.warning(f"Error checking link {url}: {str(e)}. Attempt {attempt + 1} of {max_retries}")
            if "instagram.com" in url or "reddit.com" in url or "forum.defcon.org" in url or "eventbrite.com" in url:
                time.sleep(delay)
            if attempt == max_retries - 1:
                logger.error(f"Failed to check link after {max_retries} attempts: {url}")
                return ('error', url)
    return ('error', url)

def extract_urls(text):
    urls = re.findall(r'\((https?://[^\s)]+)\)', text)
    if not urls:
        urls = re.findall(r'https?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\\(\\),]|(?:%[0-9a-fA-F][0-9a-fA-F]))+', text)
    return urls

def format_duplicate_groups(df):
    duplicates = df[df.duplicated(subset=['DCG Name'], keep=False)]
    duplicate_info = defaultdict(list)
    for _, row in duplicates.iterrows():
        duplicate_info[row['DCG Name']].append(row['Group Location'])
    formatted_duplicates = []
    for group, locations in duplicate_info.items():
        if len(locations) > 1:
            formatted_duplicates.append(f"Duplicate group: {group} in {locations[0]} and {locations[1]}")
        else:
            formatted_duplicates.append(f"Duplicate group: {group} in {locations[0]} (location repeated)")
    return formatted_duplicates

def process_xlsx(file_path):
    try:
        logger.info(f"Reading Excel file: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        data = ws.values
        columns = next(data)
        df = pd.DataFrame(data, columns=columns)
        issues = []
        redirects = []
        broken_links = []
        logger.info("Checking for duplicate groups")
        duplicate_groups = format_duplicate_groups(df)
        issues.extend(duplicate_groups)
        logger.info("Checking for broken links and other issues")
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_url = {}
            for _, row in df.iterrows():
                for column in ['Website', 'Social Link / Point of Contact (POC)', 'Join Group']:
                    content = str(row[column])
                    urls = extract_urls(content)
                    for url in urls:
                        if is_valid_url(url):
                            future_to_url[executor.submit(check_link, url)] = (row['DCG Name'], column, url)
            for future in as_completed(future_to_url):
                group_name, link_type, url = future_to_url[future]
                try:
                    result, final_url = future.result()
                    if result == 'broken':
                        issues.append(f"Broken {link_type} link for {group_name}: {url}")
                        broken_links.append((group_name, link_type, url))
                    elif result == 'redirect':
                        redirects.append((group_name, link_type, url, final_url))
                    elif result == 'error':
                        logger.warning(f"Error checking {link_type} link for {group_name}: {url}")
                except Exception as exc:
                    logger.error(f"Unexpected error checking {link_type} link for {group_name}: {url} - {exc}")
        
        for group_name, link_type, url in broken_links:
            df.loc[df['DCG Name'] == group_name, link_type] = df.loc[df['DCG Name'] == group_name, link_type].apply(lambda x: x.replace(url, '') if isinstance(x, str) else x)
        
        duplicate_groups = df[df.duplicated(subset=['DCG Name'], keep=False)]
        for _, group in duplicate_groups.groupby('DCG Name'):
            if len(group) > 1:
                df.loc[group.index[1:], 'DCG Name'] = group['DCG Name'] + '.2'
        
        return df, issues, redirects
    except PermissionError:
        logger.error(f"Error: Unable to access the input Excel file: {file_path}")
        return None, None, None
    except Exception as e:
        logger.error(f"Unexpected error processing Excel file: {str(e)}")
        return None, None, None

def save_to_excel(df, redirects, output_file):
    try:
        logger.info(f"Saving corrected XLSX file: {output_file}")
        wb = openpyxl.Workbook()
        ws = wb.active
        for col_num, column_title in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = column_title
            cell.font = openpyxl.styles.Font(bold=True)
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                if isinstance(cell_value, str):
                    cell_value = cell_value.replace('twitter.com', 'x.com')
                    cell_value = cell_value.replace('Twitter', 'X')
                cell.value = cell_value
                cell.alignment = Alignment(wrapText=True, vertical='top')
        for redirect in redirects:
            group_name, link_type, old_url, new_url = redirect
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if row[0].value == group_name:
                    for cell in row:
                        if isinstance(cell.value, str) and old_url in cell.value:
                            cell.value = cell.value.replace(old_url, new_url)
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        for row in ws.rows:
            max_height = 0
            for cell in row:
                if cell.value:
                    text_lines = str(cell.value).count('\n') + 1
                    text_height = text_lines * 15
                    max_height = max(max_height, text_height)
            ws.row_dimensions[row[0].row].height = max_height
        wb.save(output_file)
        logger.info(f"Corrected XLSX saved as: {output_file}")
    except Exception as e:
        logger.error(f"Error saving corrected XLSX file: {str(e)}")

def generate_pdf_report(issues, output_file):
    logger.info(f"Generating PDF report: {output_file}")
    doc = SimpleDocTemplate(output_file, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    content = []
    link_style = ParagraphStyle(
        'LinkStyle',
        parent=styles['BodyText'],
        textColor=colors.blue,
        underline=True,
    )
    content.append(Paragraph("DEF CON Groups Link Checker Report", styles['Title']))
    content.append(Spacer(1, 0.25*inch))
    content.append(Paragraph(f"Total issues found: {len(issues)}", styles['Heading2']))
    content.append(Spacer(1, 0.25*inch))
    data = [["Issue"]]
    for issue in issues:
        url_match = re.search(r'(https?://\S+)', issue)
        if url_match:
            url = url_match.group(1)
            link_text = f'<a href="{url}" color="blue">{url}</a>'
            formatted_issue = issue.replace(url, link_text)
            wrapped_issue = Paragraph(formatted_issue, link_style)
        else:
            wrapped_issue = Paragraph(issue, styles['BodyText'])
        data.append([wrapped_issue])
    table = Table(data, colWidths=[9.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    content.append(table)
    doc.build(content)
    logger.info(f"PDF report generated: {output_file}")

def create_output_folder():
    base_folder = "DCGroups_Reports"
    if not os.path.exists(base_folder):
        os.mkdir(base_folder)
    current_date = datetime.now().strftime("%Y%m%d")
    folder_name = f"{base_folder}/Report_{current_date}"
    counter = 1
    while os.path.exists(folder_name):
        folder_name = f"{base_folder}/Report_{current_date}_{counter}"
        counter += 1
    os.mkdir(folder_name)
    logger.info(f"Created output folder: {folder_name}")
    return folder_name

if __name__ == "__main__":
    input_file = "DCGroups-final.xlsx"
    output_folder = create_output_folder()
    output_xlsx = f"{output_folder}/DCGroups-corrected.xlsx"
    output_pdf = f"{output_folder}/DCGroups-report.pdf"
    output_chart = f"{output_folder}/DCGroups-chart.png"
    logger.info(f"Starting processing of {input_file}")
    df, issues, redirects = process_xlsx(input_file)
    if df is None or issues is None or redirects is None:
        logger.error("Script execution stopped due to file access error.")
        sys.exit(1)
    generate_pdf_report(issues, output_pdf)
    save_to_excel(df, redirects, output_xlsx)
    logger.info("Summary of issues found:")
    for issue in issues:
        logger.info(issue)
    logger.info("Summary of redirects (not included in PDF report):")
    for redirect in redirects:
        logger.info(f"Redirected link for {redirect[0]} ({redirect[1]}): {redirect[2]} -> {redirect[3]}")
    logger.info(f"Total issues found: {len(issues)}")
    logger.info(f"Total redirects found: {len(redirects)}")
    issue_types = ['Duplicate Groups', 'Broken Links']
    issue_counts = [
        sum('Duplicate group' in issue for issue in issues),
        sum('Broken' in issue for issue in issues)
    ]
    logger.info("Generating summary chart")
    plt.figure(figsize=(10, 6))
    plt.bar(issue_types, issue_counts)
    plt.title('DEF CON Groups Link Checker Results')
    plt.xlabel('Issue Type')
    plt.ylabel('Number of Issues')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(output_chart)
    logger.info(f"Chart saved as: {output_chart}")
    logger.info("Script execution completed successfully")